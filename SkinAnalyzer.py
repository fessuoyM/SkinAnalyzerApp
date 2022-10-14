#Copyright (c) 2022 Youssef Mohamed

import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from PIL import Image, ImageTk
import os
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from skimage import color
import math

## Some Constants to be used globally
BACKGROUND = '#225235'#'#222222' 348253
BUTTON_BACKGROUND = '#CDC399'#"#888888" 205195153
photoFolder = None
excelFile = None
data = {}
numSamples = 6
maxSamples = 20
minSamples = 1
maxPicsPerCol=4
ind = -1
width = 10
blank = Image.fromarray(np.repeat(np.repeat(np.array([[[255,255,255]]], np.uint8),width*2,axis=1),width*2,axis=0))
mode = None
fileList = None
maxDim = (600, 300)
scopeMag = 4
coords = []
scoreThresholds = np.array([55, 41, 28, 10, -30])
imgMain = None
book = None
alreadyDone = []
ws = None
name = None
flagText = False
flagFolder = False
prog = None
style = None
pbar = None

## Initial setup of the window
root = Tk()
root.title('Skin Analyzer')
w  = root.winfo_screenwidth()
h = root.winfo_screenheight()
root.geometry(f'{int(w/2)}x{int(h*0.8)}')
root.configure(background=BACKGROUND)
root.bind("<Escape>", lambda e: e.widget.quit())
setupFrame = Frame(root, background=BUTTON_BACKGROUND)
setupFrame.pack(fill='none', expand=True)

## Setup of the main screen before data collection
def mainSetup():
    global multiPointBut, burnPhotoBut, flagText, flagFolder, photoFolder, name, numSamples
    #Frames for orginazation
    fTop = Frame(setupFrame, background=BUTTON_BACKGROUND)
    fMid = Frame(setupFrame, background=BUTTON_BACKGROUND)
    fBottom = Frame(setupFrame, background=BUTTON_BACKGROUND)
    f1 = Frame(fMid, padx=10, background=BUTTON_BACKGROUND)
    f2 = Frame(fMid, padx=10, background=BUTTON_BACKGROUND)

    #Labels for button groups
    l_left = Label(fTop, text="Select the folder containg the images and the excel file for the ouptut.", background=BUTTON_BACKGROUND, wraplength=245, justify='center')
    l_left.pack(side='left', fill = 'both', padx = (36,0))
    l_right = Label(fTop, text="Select the type of collection you will be using.", background=BUTTON_BACKGROUND, wraplength=315, justify='center')
    l_right.pack(side='right', fill='both', padx = (0,48))

    #Buttons to choose data collection method
    fMulti = Frame(f1, background=BUTTON_BACKGROUND)
    multiPointBut = Button(fMulti, text="MultiPoint Collection", highlightbackground=BUTTON_BACKGROUND, height=2, state=tk.DISABLED, command=lambda:setupCollection('mp'))
    multiPointBut.pack(side='left', fill='x', expand=True)
    e2 = Entry(fMulti, text='', width = 2, validate = "focusout", validatecommand=lambda:checkNumeric(e2, False))
    e2.insert(END, str(numSamples))
    e2.pack(side='right')
    multiLabel = Label(fMulti, text = "Samples/Pic: ", width = 10, height=2, background=BUTTON_BACKGROUND)
    multiLabel.pack(side='right', padx=(6,0))

    burnPhotoBut = Button(f1, text="Two-point with ΔE", highlightbackground=BUTTON_BACKGROUND, width=35, height=2 ,state=tk.DISABLED, command=lambda:setupCollection('bp'))

    if flagText and flagFolder and not photoFolder=='':
        multiPointBut['state']=tk.NORMAL
        burnPhotoBut['state']=tk.NORMAL

    #Buttons to get file paths
    getPhotoFolder = Button(f2, text="Select Photos Folder", highlightbackground=BUTTON_BACKGROUND, width=25, height=2, command=lambda: getPath("Select Photos Folder", True))
    getExcelFile = Button(f2, text="Select Excel File", highlightbackground=BUTTON_BACKGROUND, width=25, height=2, command=lambda: getPath("Select Excel File", False))

    #Text entry to get names
    Label(fBottom, text = "Name: ", width = 10, height=2, background=BUTTON_BACKGROUND).pack(side='left', padx=(70,0))
    e = Entry(fBottom, text='', width = 20, validate = "focusout", validatecommand=lambda:getName(e))
    if name is not None:
        e.insert(END, name)
    e.pack(side='left', padx=(0,70))
    fBottom.bind("<Button 1>", lambda e: setupFrame.focus_set())
    setupFrame.bind("<Button 1>", lambda e: setupFrame.focus_set())
    getPhotoFolder.bind("<Button 1>", lambda e: setupFrame.focus_set())
    getExcelFile.bind("<Button 1>", lambda e: setupFrame.focus_set())
    multiLabel.bind("<Button 1>", lambda e: setupFrame.focus_set())
    fMulti.bind("<Button 1>", lambda e: setupFrame.focus_set())
    l_left.bind("<Button 1>", lambda e: setupFrame.focus_set())
    l_right.bind("<Button 1>", lambda e: setupFrame.focus_set())
    root.bind("<Return>", lambda e: setupFrame.focus_set())

    #Text entry for apperture
    e1 = Entry(fBottom, text='', width = 3, validate = "focusout", validatecommand=lambda:checkNumeric(e1, True))
    e1.insert(END, str(width))
    e1.pack(side='right', padx=(0,70))
    Label(fBottom, text = "Aperture: ", width = 10, height=2, background=BUTTON_BACKGROUND).pack(side='right')


    #Label with directions
    l1 = Label(setupFrame, text="Welcome to Skin Analyzer, you must enter your name and choose a folder containing the pictures to be analyzed to proceed. Optionally, you may select a an excel file from a previous session (Must be the same format, i.e. you clicked the same button on the right last time). Finally select a analysis mode from the right to get started.", wraplength=550, justify='center', background=BUTTON_BACKGROUND)
    l1.pack(side='top', pady=(20,10), padx=10)
    fBottom.pack(side='top', pady=(10,0))
    l1.bind("<Button 1>", lambda e: setupFrame.focus_set())

    #Label with ownership
    meetTheTeam = Button(setupFrame, text = "Meet the rest of the team here", highlightbackground=BUTTON_BACKGROUND, width=40, height=2, command=meetTheTeamPage)
    meetTheTeam.pack(side='bottom', pady=(0,10))
    l3 = Label(setupFrame, text="Youssef Mohamed, PHD and Bilal Koussayer, BS", wraplength = 500, justify='center', background=BUTTON_BACKGROUND)
    l3.pack(side='bottom', pady = (0,10))
    l2 = Label(setupFrame, text="This Software was developed by the Morsani College Of Medicine and Tampa General Hospital Plastic Surgery Burn Research Team lead by", wraplength = 500, justify='center', background=BUTTON_BACKGROUND)
    l2.pack(side='bottom', pady = (10,0))



    #Pack the components onto the setupFrame
    fMulti.pack(fill='x', expand=True)
    burnPhotoBut.pack()
    getPhotoFolder.pack()
    getExcelFile.pack()
    f1.pack(side='right')
    f2.pack(side='left')
    fTop.pack(pady=(20,0), fill='x')
    fMid.pack(pady=(0,20))


## Check numeric input for Apperture
def checkNumeric(e, app):
    global width, numSamples
    if e.get().isdigit():
        if app:
            width = int(e.get())
            if width > 100:
                width = 100
            elif width < 0:
                width = 0
            e.delete(0,END)
            e.insert(END,str(width))
        else:
            numSamples = int(e.get())
            if numSamples > maxSamples:
                numSamples = maxSamples
            elif numSamples < minSamples:
                numSamples = minSamples
            e.delete(0,END)
            e.insert(END,str(numSamples))
        return True
    else:
        e.delete(0,END)
        if app:
            e.insert(END,str(width))
        else:
            e.insert(END,str(numSamples))
        return False

## Function wrappers to obtain the file paths needed as requested by buttons
def getPath(str, i):
    global photoFolder, excelFile, book, alreadyDone, ws, flagText, flagFolder
    if i:
        photoFolder = filedialog.askdirectory(title=str)
        if not photoFolder=='':
            if flagText:
                multiPointBut['state']=tk.NORMAL
                burnPhotoBut['state']=tk.NORMAL
            flagFolder = True

    else:
        excelFile = filedialog.askopenfilename(title=str, filetypes=(("Excel File", "*.xlsx"),('All files', '*.*')))
        if excelFile == '':
            excelFile = None
        book = load_workbook(excelFile)
        ws=book.active
        col=ws['A']
        alreadyDone = list(set([c.value for c in col]))
        print(excelFile)

## Get the name of the user
def getName(e):
    global name, flagText, flagFolder
    name = e.get()
    print(f'Hey {name}')
    if flagFolder:
        multiPointBut['state']=tk.NORMAL
        burnPhotoBut['state']=tk.NORMAL
    flagText = True

## Setup and run the photoValidation program
def setupCollection(m):
    global setupFrame, numSamples, mode, photoFolder, fileList, alreadyDone, root, data, name
    mode = m
    setupFrame.destroy()
    print(mode)
    if mode == "mp":
        txt = ["Sample #{0}".format(x) for x in range(1,numSamples+1)]
        data = {'filename': [], 'imageNum': [], 'name': [], 'Point': [], 'x_corr': [], 'y_corr': [], 'L*': [], 'a*': [], 'b*': [], 'ITA': [], 'Fitzpatrick Skin Type': [], 'R': [], 'G': [], 'B': []}
    elif mode == "bp":
        numSamples = 2
        txt = ["Skin","Burn"]
        data = {'filename': [], 'imageNum': [], 'name': [], 'x_corr_Skin': [], 'y_corr_Skin': [], 'L*_Skin': [], 'a*_Skin': [], 'b*_Skin': [], 'ITA_Skin': [], 'Fitzpatrick Skin Type_Skin': [], 'x_corr_Burn': [], 'y_corr_Burn': [], 'L*_Burn': [], 'a*_Burn': [], 'b*_Burn': [], 'ITA_Burn': [], 'Fitzpatrick Skin Type_Burn': [], 'DeltaE': [], 'R_Skin': [], 'G_Skin': [], 'B_Skin': [], 'R_Burn': [], 'G_Burn': [], 'B_Burn': []}
    else:
        numSamples = 0
        txt=""
    print(f"numSamples={numSamples}")
    fileList = os.listdir(photoFolder)
    fileList = [f for f in fileList if ('.png' in f or '.jpeg' in f or '.jpg' in f) and f not in alreadyDone]
    fileList.sort()
    if len(fileList)==0:
        Label(root, text="You have already analyzed all the photos in this folder", width = 50, height = 2, background=BUTTON_BACKGROUND).pack()
        root.after(3000, lambda: root.quit())
    setupCollectionFrame(numSamples, txt)

## Setup the frame for data collection as a function of number of samples to take
def setupCollectionFrame(n,txt):
    global root, ind, fileList, maxDim, scope, canvs, scopeMag, imgMain, prog, style, pbar, maxPicsPerCol, numSamples
    f0 = Frame(root, background=BACKGROUND)
    f1 = Frame(f0, background=BACKGROUND)
    f2 = Frame(f0, background=BACKGROUND)
    ind=0
    #Get main picture plotted
    print(os.path.join(photoFolder,fileList[ind]))
    imgMain = Image.open(os.path.join(photoFolder,fileList[ind]))
    print(maxDim)
    maxDim = (root.winfo_width()-math.ceil(numSamples/maxPicsPerCol)*100, root.winfo_height()-120)
    dim = int(imgMain.size[0]/maxDim[0] < imgMain.size[1]/maxDim[1])
    pixels_x, pixels_y = tuple([int(maxDim[dim]/imgMain.size[dim] * x)  for x in imgMain.size])
    imgb = ImageTk.PhotoImage(imgMain.resize((pixels_x, pixels_y)))
    mainImg = Label(f1, image = imgb, borderwidth = 0)
    mainImg.image = imgb
    mainImg.pack(fill='both')
    mainImg.bind("<Button 1>", lambda e: mainImgClick(e, samples, scaler))
    mainImg.bind("<Motion>", lambda e: updateScope(e,f1,scope, canvs,scaler, width))
    f1.pack(fill='both',side='left')


    #Create the Scope
    scaler = int(50/blank.size[0])*blank.size[0]
    f3 = Frame(f2, background=BACKGROUND)
    imgTemp=ImageTk.PhotoImage(blank.resize((scaler,scaler)))
    canvs = Canvas(f3, width=scaler*2-3, height=scaler*2-3, background=BACKGROUND)
    scope = canvs.create_image(0,0,anchor='nw',image=imgTemp)
    Label(f2, text = "Select Tool", background=BUTTON_BACKGROUND, width=12).pack(pady=(10,0), padx=(0,18))
    canvs.pack(pady=(0,20), side='left')

    #Add box to scope
    line=[]
    line.append(canvs.create_line((scaler*(1-1/scopeMag), scaler*(1-1/scopeMag), scaler*(1-1/scopeMag), scaler*(1+1/scopeMag)), fill='black', width=2))
    line.append(canvs.create_line((scaler*(1-1/scopeMag), scaler*(1-1/scopeMag), scaler*(1+1/scopeMag), scaler*(1-1/scopeMag)), fill='black', width=2))
    line.append(canvs.create_line((scaler*(1+1/scopeMag), scaler*(1+1/scopeMag), scaler*(1-1/scopeMag), scaler*(1+1/scopeMag)), fill='black', width=2))
    line.append(canvs.create_line((scaler*(1+1/scopeMag), scaler*(1+1/scopeMag), scaler*(1+1/scopeMag), scaler*(1-1/scopeMag)), fill='black', width=2))


    #Create the zoom buttons
    f4 = Frame(f3, background=BACKGROUND)
    Button(f4, text="+", highlightbackground=BUTTON_BACKGROUND, command=lambda:updateZoom(0.75, scaler, line)).pack(pady=(2,1))
    Button(f4, text="-", highlightbackground=BUTTON_BACKGROUND, command=lambda:updateZoom(1.5, scaler, line)).pack(pady=(1,2))
    f4.pack(side='right')
    f3.pack(pady=(0,20))

    #Get samples plotter on right hand side
    samples=[]
    fhorz = None
    fsamp = None
    cols = math.ceil(n/maxPicsPerCol)
    for i in range(n):
        coords.append(None)
        def make_lambda(x):
            return lambda e: redo(e, x, samples, scaler)
        imgTemp=ImageTk.PhotoImage(blank.resize((scaler,scaler)))

        if i%cols==0:
            fhorz = Frame(f2, background=BACKGROUND)
        fsamp = Frame(fhorz, background=BACKGROUND)
        samples.append(Label(fsamp,image=imgTemp))
        Label(fsamp, text = txt[i], background=BUTTON_BACKGROUND, width=8).pack(pady=(10,0), padx=2)
        samples[i].pack(pady=(0,10), padx=(8,0))
        samples[i].image = imgTemp
        samples[i].bind("<Button 1>", make_lambda(i))
        fsamp.pack(side='left')
        if ((i+1)%cols==0) or (i+1==n):
            fhorz.pack()

    #next and finish button
    f5 = Frame(f2, background=BACKGROUND)
    next = Button(f5, text="Next", width=8, height=2, highlightbackground=BUTTON_BACKGROUND, command=lambda:nextImg(mainImg, samples, scaler, next, f1))
    root.bind("<Return>", lambda e: nextImg(mainImg, samples, scaler, next, f5))
    next.pack(side = 'left', pady=20, padx=(0,15))
    if len(fileList)<=1:
        next['state']=tk.DISABLED
    Button(f5, text="Finish", width=8, height=2, highlightbackground="green", command=finish).pack(side = 'right', pady=20, padx=(0,15))


    f5.pack(side='right', padx=20, expand = True)
    f2.pack()

    #Label with ownership
    l3 = Label(root, text="Youssef Mohamed, PHD and Bilal Koussayer, BS", wraplength = 500, justify='center', background=BUTTON_BACKGROUND)
    l3.pack(side='bottom', fill='x')
    l2 = Label(root, text="This Software was developed by the Morsani College Of Medicine and Tampa General Hospital Plastic Surgery Burn Research Team lead by", wraplength = 500, justify='center', background=BUTTON_BACKGROUND)
    l2.pack(side='bottom', fill='x')

    #Add a progress bar
    f0.pack(fill='both', expand=True)
    style = ttk.Style(root)
    style.layout('text.Horizontal.TProgressbar',[('Horizontal.Progressbar.trough',{'children': [('Horizontal.Progressbar.pbar',{'side': 'left', 'sticky': 'ns'})],'sticky': 'nswe'}),('Horizontal.Progressbar.label', {'sticky': 'nswe'})])
    style.configure('text.Horizontal.TProgressbar', text='Progress: {0}%'.format(int(100*(ind+1)/len(fileList))), anchor='center', background='green')
    prog = DoubleVar()
    prog.set(100*(ind+1)/len(fileList))
    pbar = ttk.Progressbar(root, style='text.Horizontal.TProgressbar', length=400, mode='determinate', variable=prog)
    pbar.pack(side='bottom', pady=(12))
    pbar.step(ind/len(fileList))

    #Force update and bind reconfigure
    root.update()
    root.bind("<Configure>", lambda a: resiz(a, mainImg, f1))

##Take care of main image resize with window resize
def resiz(e,mainImg, f1):
    global maxDim, imgMain, root, maxPicsPerCol, numSamples
    root.unbind('<Configure>')
    maxDim = (e.width-math.ceil(numSamples/maxPicsPerCol)*100, e.height-120)
    dim = imgMain.size[0]/maxDim[0] < imgMain.size[1]/maxDim[1]
    pixels_x, pixels_y = tuple([int(maxDim[dim]/imgMain.size[dim] * x)  for x in imgMain.size])
    img = ImageTk.PhotoImage(imgMain.resize((pixels_x, pixels_y)))
    mainImg.configure(image=img)
    mainImg.image = img
    root.update()
    root.bind("<Configure>", lambda a: resiz(a, mainImg, f1))



## Take care of resetting after each picture
def nextImg(mainImg, samples, scaler, next, f1):
    global ind, imgMain, fileList, style, maxDim, coords, maxPicsPerCol, numSamples
    updateMasterList(imgMain)
    ind=ind+1
    prog.set(100*(ind+1)/len(fileList))
    # pbar.step(ind/len(fileList))
    pbar['value']=prog.get()
    style.configure('text.Horizontal.TProgressbar',
                    text="Progress: {0}%".format(int(prog.get())))

    #change main image to next
    root.unbind('<Configure>')
    maxDim = (root.winfo_width()-math.ceil(numSamples/maxPicsPerCol)*100, root.winfo_height()-120)
    imgMain = Image.open(os.path.join(photoFolder,fileList[ind]))
    dim = imgMain.size[0]/maxDim[0] < imgMain.size[1]/maxDim[1]
    pixels_x, pixels_y = tuple([int(maxDim[dim]/imgMain.size[dim] * x)  for x in imgMain.size])
    print((pixels_x, pixels_y))
    img = ImageTk.PhotoImage(imgMain.resize((pixels_x, pixels_y)))
    mainImg.configure(image=img)
    mainImg.image = img
    root.update()
    root.bind("<Configure>", lambda a: resiz(a, mainImg, f1))

    #rest the sample images
    for i,s in enumerate(samples):
        imgTemp=ImageTk.PhotoImage(blank.resize((scaler,scaler)))
        s.configure(image=imgTemp)
        s.image = imgTemp
        coords[i]=None

    if ind >=len(fileList)-1:
        next['state']=tk.DISABLED

## Capture the image position
def mainImgClick(eo, s, scaler):
    global width, coords, imgMain
    print("{0}, {1}".format(eo.x, eo.y))
    dim = imgMain.size[0]/maxDim[0] < imgMain.size[1]/maxDim[1]
    dim = tuple([int(maxDim[dim]/imgMain.size[dim] * x)  for x in imgMain.size])
    coord = (int(imgMain.size[0]*(eo.x-1)/dim[0]),int(imgMain.size[1]*(eo.y-1)/dim[1]))
    imgTemp=ImageTk.PhotoImage(imgMain.resize((scaler,scaler), box=(coord[0]-width,coord[1]-width,coord[0]+width,coord[1]+width)))
    sInd = [i for i,a in enumerate(coords) if a is None]
    if len(sInd)<=0:
        sInd = [len(coords)-1]
    sInd=sInd[0]
    s[sInd].configure(image=imgTemp)
    s[sInd].image = imgTemp
    coords[sInd]=coord


## Set Redo index
def redo(e,i,s,scaler):
    global coords
    imgTemp=ImageTk.PhotoImage(blank.resize((scaler,scaler)))
    s[i].configure(image=imgTemp)
    s[i].image = imgTemp
    coords[i]=None

#Update the scope window
def updateScope(eo, f1, scope, canvs, scaler, width):
    global imgMain, maxDim
    dim = imgMain.size[0]/maxDim[0] < imgMain.size[1]/maxDim[1]
    dim = tuple([int(maxDim[dim]/imgMain.size[dim] * x)  for x in imgMain.size])
    coord = [int(imgMain.size[0]*(eo.x-1)/dim[0]),int(imgMain.size[1]*(eo.y-1)/dim[1])]
    if coord[0]<width*scopeMag:
        coord[0]=width*scopeMag
    elif coord[0]+width*scopeMag>=imgMain.size[0]:
        coord[0]=imgMain.size[0]-width*scopeMag
    if coord[1]<width*scopeMag:
        coord[1]=width*scopeMag
    elif coord[1]+width*scopeMag>=imgMain.size[1]:
        coord[1]=imgMain.size[1]-width*scopeMag
    imgTemp=ImageTk.PhotoImage(imgMain.resize((scaler*2,scaler*2), box=(coord[0]-width*scopeMag,coord[1]-width*scopeMag,coord[0]+width*scopeMag,coord[1]+width*scopeMag)))
    canvs.itemconfig(scope,image=imgTemp)
    canvs.image=imgTemp

## Update scope magnification
def updateZoom(i, scaler, line):
    global scopeMag
    scopeMag*=i
    if scopeMag > 10:
        scopeMag = 10
    elif scopeMag < 2:
        scopeMag = 2
    for i in range(4):
        canvs.delete(line[i])
    line[0] = canvs.create_line((scaler*(1-1/scopeMag), scaler*(1-1/scopeMag), scaler*(1-1/scopeMag), scaler*(1+1/scopeMag)), fill='black', width=2)
    line[1] = canvs.create_line((scaler*(1-1/scopeMag), scaler*(1-1/scopeMag), scaler*(1+1/scopeMag), scaler*(1-1/scopeMag)), fill='black', width=2)
    line[2] = canvs.create_line((scaler*(1+1/scopeMag), scaler*(1+1/scopeMag), scaler*(1-1/scopeMag), scaler*(1+1/scopeMag)), fill='black', width=2)
    line[3] = canvs.create_line((scaler*(1+1/scopeMag), scaler*(1+1/scopeMag), scaler*(1+1/scopeMag), scaler*(1-1/scopeMag)), fill='black', width=2)

## Update the samples into master list
def updateMasterList(im):
    global coords, ind, fileList,data, name
    im = np.array(im)[:,:,:3]
    if mode == 'mp':
        for i, coord in enumerate(coords):
            if coord is None:
                meanRGB = [-9999,-9999,-9999]
                meanLAB = meanRGB
                coord=[-9999,-9999]
                score=-9999
            else:
                RGB = im[int(coord[1])-width:int(coord[1])+width+1,int(coord[0])-width:int(coord[0])+width+1,0:3]
                LAB = color.rgb2lab(RGB)
                meanRGB=np.mean(RGB,axis=(0,1))
                meanLAB=np.mean(LAB,axis=(0,1))
                score=math.atan2(meanLAB[0]-50,meanLAB[2])*180/math.pi
            data['filename'].append(fileList[ind])
            data['imageNum'].append(ind)
            data['name']=name
            data['Point'].append(i+1)
            data['x_corr'].append(int(coord[0]))
            data['y_corr'].append(int(coord[1]))
            data['R'].append(meanRGB[0])
            data['G'].append(meanRGB[1])
            data['B'].append(meanRGB[2])
            data['L*'].append(meanLAB[0])
            data['a*'].append(meanLAB[1])
            data['b*'].append(meanLAB[2])
            data['ITA'].append(score)
            if coord[0]<0:
                data['Fitzpatrick Skin Type'].append(-9999)
            else:
                data['Fitzpatrick Skin Type'].append(sum(score<scoreThresholds)+1)
    elif mode == 'bp':
        RGB = im[int(coords[0][1])-width:int(coords[0][1])+width+1,int(coords[0][0])-width:int(coords[0][0])+width+1,0:3]
        LAB = color.rgb2lab(RGB)
        data['filename'].append(fileList[ind])
        data['imageNum'].append(ind)
        data['name']=name
        data['x_corr_Skin'].append(int(coords[0][0]))
        data['y_corr_Skin'].append(int(coords[0][1]))
        meanRGB=np.mean(RGB,axis=(0,1))
        data['R_Skin'].append(meanRGB[0])
        data['G_Skin'].append(meanRGB[1])
        data['B_Skin'].append(meanRGB[2])
        meanLAB=np.mean(LAB,axis=(0,1))
        data['L*_Skin'].append(meanLAB[0])
        data['a*_Skin'].append(meanLAB[1])
        data['b*_Skin'].append(meanLAB[2])
        meanLAB1 = meanLAB
        score=math.atan2(meanLAB[0]-50,meanLAB[2])*180/math.pi
        data['ITA_Skin'].append(score)
        data['Fitzpatrick Skin Type_Skin'].append(sum(score<scoreThresholds)+1)
        RGB = im[int(coords[1][1])-width:int(coords[1][1])+width+1,int(coords[1][0])-width:int(coords[1][0])+width+1,0:3]
        LAB = color.rgb2lab(RGB)
        data['x_corr_Burn'].append(int(coords[1][0]))
        data['y_corr_Burn'].append(int(coords[1][1]))
        meanRGB=np.mean(RGB,axis=(0,1))
        data['R_Burn'].append(meanRGB[0])
        data['G_Burn'].append(meanRGB[1])
        data['B_Burn'].append(meanRGB[2])
        meanLAB=np.mean(LAB,axis=(0,1))
        data['L*_Burn'].append(meanLAB[0])
        data['a*_Burn'].append(meanLAB[1])
        data['b*_Burn'].append(meanLAB[2])
        score=math.atan2(meanLAB[0]-50,meanLAB[2])*180/math.pi
        data['ITA_Burn'].append(score)
        data['Fitzpatrick Skin Type_Burn'].append(sum(score<scoreThresholds)+1)
        data['DeltaE'].append(math.sqrt(math.pow(meanLAB1[0]-meanLAB[0],2)+math.pow(meanLAB1[1]-meanLAB[1],2)+math.pow(meanLAB1[2]-meanLAB[2],2)))
    print(data)

## Finish up and paste into an excel file
def finish():
    global root, book, imgMain, ws, mode, excelFile, coords
    updateMasterList(imgMain)
    pdData = pd.DataFrame(data)
    pdData = pdData.replace(-9999,"")
    # pdData.sort_values(['filename','Point'], axis=0, inplace=True, na_position='last')
    print(pdData)
    if excelFile is None:
        if mode == "mp":
            f="PhotoVaildationData.xlsx"
        elif mode == "sv":
            f="SkinVaildationData.xlsx"
        elif mode == "bp":
            f="DataCollection.xlsx"
        else:
            f="UknownMode.xlsx"

        if os.path.exists(os.path.join(photoFolder,f)):
            i=1
            while os.path.exists(os.path.join(photoFolder,"{0}({1}).{2}".format(f.split(".")[0],i,f.split(".")[1]))):
                i+=1
            f="{0}({1}).{2}".format(f.split(".")[0],i,f.split(".")[1])

        xlWriter = pd.ExcelWriter(os.path.join(photoFolder,f))
        if not mode=='sv':
            pdData.to_excel(xlWriter,index=False)
        else:
            pdData.filter(regex='name').to_excel(xlWriter,index=False)
            pdData.filter(regex='Fitzpatrick Skin Type').to_excel(xlWriter,index=False, startrow=0, startcol=1)
            pdData.drop(pdData.filter(regex='name|Fitzpatrick Skin Type').columns,axis=1).to_excel(xlWriter,index=False, startrow=0,  startcol=1+len(pdData.filter(regex='Fitzpatrick Skin Type').columns))
        xlWriter.save()
    else:
        xlWriter = pd.ExcelWriter(excelFile, engine="openpyxl")
        xlWriter.book = book
        xlWriter.sheets = {ws.title: ws for ws in book.worksheets}
        if not mode=='sv':
            data['imageNum']=list(np.array(data['imageNum'])+ws['B'][-1].value+1)
            pdData = pd.DataFrame(data)
            pdData.to_excel(xlWriter,index=False,header=False,startrow=xlWriter.sheets['Sheet1'].max_row)
        else:
            row=xlWriter.sheets['Sheet1'].max_row
            pdData.filter(regex='name').to_excel(xlWriter,index=False,header=False,startrow=row)
            pdData.filter(regex='Fitzpatrick Skin Type').to_excel(xlWriter,index=False,header=False,startrow=row, startcol=1)
            pdData.drop(pdData.filter(regex='name|Fitzpatrick Skin Type').columns,axis=1).to_excel(xlWriter,index=False,header=False,startrow=row, startcol=1+len(pdData.filter(regex='Fitzpatrick Skin Type').columns))
        xlWriter.save()
    coords = []
    root.quit()

## Show the contributors tot he project
def meetTheTeamPage():
    global setupFrame
    setupFrame.destroy()
    MTT = Frame(root, background=BUTTON_BACKGROUND)
    MTT.pack(fill='none', expand=True)

    l2 = Label(MTT, text="This Software was developed by the Morsani College Of Medicine and Tampa General Hospital Plastic Surgery Burn Research Team lead by", wraplength = 500, justify='center', background=BUTTON_BACKGROUND)
    l2.pack(side='top', pady = (10,0))
    l3 = Label(MTT, text="Youssef Mohamed, PHD and Bilal Koussayer, BS", wraplength = 500, justify='center', background=BUTTON_BACKGROUND)
    l3.pack(side='top', pady = (0,10))

    colFrame = Frame(MTT, background=BUTTON_BACKGROUND)
    colFrame.pack(side='top', pady=(10,0), fill='none', expand=True)
    col1Frame = Frame(colFrame, background=BUTTON_BACKGROUND)
    col2Frame = Frame(colFrame, background=BUTTON_BACKGROUND)
    col1Frame.pack(side='left', padx=(20,20), anchor='ne', fill='x')
    col2Frame.pack(side='right', padx=(20,20), anchor='nw', fill='x')

    l0_1 = Label(col1Frame, text="TGH Team:", width=13, height=2, font= 'bold 15', background=BUTTON_BACKGROUND)
    l1_1 = Label(col2Frame, text="Medical Students\nTeam:", width=13, height=2, font='bold 15', background=BUTTON_BACKGROUND)

    l0 = Label(col1Frame, text="Jake Laun, MD\nNicole Le, MD\nKristen Whalen, MD\nKristina Gemayel, DO\nMahmood Al Bayati, MD\nLoryn Taylor, ARNP", width=16, background=BUTTON_BACKGROUND)
    l1 = Label(col2Frame, text="Ellie Randolph, BS\nJaynie Criscione, BS\nJulia Morris, BS\nMarian Mikhael, BS\nRithvic Jupudi, BS\nSarah Moffitt, BS\nShreya Arora, BS\nTimothy Nehila,BS\nWilliam West III, BS\nM. Tahseen Alkaelani, BS\nAdam Mohamed, MS", width=16, background=BUTTON_BACKGROUND)

    l0_1.pack(side='top', padx=(20,20), fill='x')
    l1_1.pack(side='top', padx=(20,20), fill='x')
    l0.pack(side='bottom', padx=(20,20), fill='x')
    l1.pack(side='bottom', padx=(20,20), fill='x')

    l4 = Label(MTT, text="Please use the reference below if you use this software in your research:\n\nMohamed, Y., Koussayer, B., Le, N., Whalen, K., Al Bayati, M., Gemayel, K., Taylor, L., Randolph, E., Criscione, J., Morris, J., Mikhael, M., Jupudi, R., Moffitt, S., Arora, S., Nehila, T., West, W., & Laun, J. (1010, October 10). Skin AnalyzerVersion (1). WEBSITE_TITLE_HERE. Retrieved from WEBSITE_URKL_HERE.", wraplength = 500, justify='center', background=BUTTON_BACKGROUND)


    back = Button(MTT, text="Back to Main Page", width = 20, height=2, background=BUTTON_BACKGROUND, highlightbackground=BUTTON_BACKGROUND, command=lambda: setupMainAgain(MTT))
    back.pack(side='bottom', pady=(20,10))
    l4.pack(side='bottom', pady=(10,5), padx = 30)

## Setup the setupFrame again
def setupMainAgain(MTT):
    global setupFrame, root
    MTT.destroy()
    w  = root.winfo_screenwidth()
    h = root.winfo_screenheight()
    root.geometry(f'{int(w/2)}x{int(h*0.8)}')
    root.configure(background=BACKGROUND)
    setupFrame = Frame(root, background=BUTTON_BACKGROUND)
    setupFrame.pack(fill='none', expand=True)
    mainSetup()

## run the setup function and start the event loop
mainSetup()
root.mainloop()
